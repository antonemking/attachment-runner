###############################
# Antone M King               #
# Push Attachments to Recods  #
###############################

import requests
import shutil
import zipfile
import time
import os
from pathlib import Path
import config
import getType
import pprint
import json
from openpyxl import load_workbook, Workbook
import sys
from texttable import Texttable

t = Texttable()
user = config.username
pwd = config.password

url = config.TARGET_INSTANCE + '/api/now/attachment/upload'
table = config.table
file_name = config.ATTACHMENT_LOG_PATH

wb = load_workbook(file_name)
ws = wb['Sheet']

log_headers = ['Correlation Sys ID', 'Original Number', 'File Name', 'Result']
log_book = Workbook()
log_sheet = log_book.active
log_sheet.append(log_headers)

root = config.ATTACHMENT_EXTRACTION_PATH
fail_path = config.FAILED_ATTACHMENT_PATH
file_count = 0
failed = 0

def getSysID(table, correField, correlationId):
    url = config.TARGET_INSTANCE + config.ATTACHMENT_RESOURCE + '/table/' +table+ '/correlationfield/' + correField+ '/correlationID/' + correlationId
    headers = {"Content-Type": "application/json", "Accept": "application/json"}
    r = requests.get(url, auth=(user, pwd), headers=headers, verify=False)
    if r.status_code != 200: 
        print('Status:', r.status_code, 'Headers:', r.headers, 'Error Response:',r.json())
        exit()
    data = json.loads(r.text)
    result = str(data['result'][0]['sys_id'])
    
    return result

for row in ws.iter_rows(min_row=2):
    try:
        correlation_sys_id = row[0].value
        record_number = row[1].value
        sys_id = getSysID(table, config.CORRELATION_FIELD_NAME, correlation_sys_id)    
        print("================================")
        print("")     
        print("Table Sys Id " + sys_id)
        print("Correlation Sys ID " + correlation_sys_id)
        print("Original Record Number " + record_number)
        directory =  root + record_number
        fail_dir = fail_path + record_number

        print("Directory " + directory) 
        for r, dirs, filenames in os.walk(directory):
            for f in filenames:
                        
                contenttype = getType.getContentType(f)
            
                #print("File Extension " + file_extension)
                #print("Content-Type " + contenttype)
                
                payload = {'table_name': table, 'table_sys_id': sys_id}
                
                file_path = directory + '/' + f
                with open(file_path, 'rb') as doc:
                    files = {'file': (file_path , doc, contenttype , {'Expires': '0'} )}

                    headers = {"Accept":"*/*"}

                    r = requests.post(url, auth=(user, pwd), headers=headers, files=files, data=payload, verify=False)

                    # Check for HTTP codes other than 201
                    if r.status_code is 201:
                        file_count += 1
                        log_sheet.append([ correlation_sys_id, record_number, f, 'Success - ' + str(r.status_code)])
                        print("Uploading file -----> " + f)
                        
                    else:
                        print("")
                        print("Status Code " + str(r.status_code) + " for file ---> " + f )
                        print("")
                        print("Moving File " + f + " moved to failed directoy --> " + fail_dir)
                        failed +=1
                        log_sheet.append([ correlation_sys_id, record_number, f, 'Failed to Upload  - ' + str(r.status_code)])
                        if os.path.isdir(fail_dir):
                            shutil.copy(file_path, fail_dir)
                        else:
                            os.makedirs(fail_dir)
                            shutil.copy(file_path, fail_dir)
                       
           
        print("")
    except Exception as e:
        log_sheet.append([ correlation_sys_id, record_number, f, 'Error - ' + 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno) + ' ' + str(e)])

      
log_book.save('upload-log.xlsx')
print("================================")
print("")
print("DONE! ")
print("")

t.add_rows([['Uploaded', 'Failed'], [str(file_count), str(failed)]])

print(t.draw())
print("")
print("Please view the upload-log file for additional information")
