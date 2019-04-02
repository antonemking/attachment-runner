###############################
# Antone M King               #
# Convert bad file extensions #
###############################


import time
import os, sys
from pathlib import Path
import config
from openpyxl import load_workbook, Workbook
from texttable import Texttable

t = Texttable()
table = config.table
file_name = config.ATTACHMENT_LOG_PATH

wb = load_workbook(file_name)
ws = wb['Sheet']

root = config.ATTACHMENT_EXTRACTION_PATH
converted= 0

for row in ws.iter_rows(min_row=2):
    try:
        
        record_number = row[1].value
       
        print("================================")
             
        print("Original Record Number " + record_number)
        directory =  root + record_number
       

        print("Directory " + directory) 
        for r, dirs, filenames in os.walk(directory):
            for f in filenames:
                if f.lower().endswith('.pdf_undefined'):
                    infilename = os.path.join(directory,f)
                    newname = infilename.replace('.pdf_undefined', '.pdf')
                    output = os.rename(infilename, newname)
                    converted+=1    
        print("")
    except Exception as e:
        print(str(e))

print("================================")
print("")
print("DONE! ")
print("")

t.add_rows([['Converted'], [str(converted)]])

print(t.draw())


