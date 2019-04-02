#####################################
# Antone M King                     #
# Zip and save attachments locally  #
#####################################

import requests
import zipfile
import io
import config
from openpyxl import load_workbook, Workbook

file_name = config.OUTPUT_FILE_PATH
extract_path = config.ATTACHMENT_EXTRACTION_PATH
table = config.table

wb = load_workbook(file_name)
ws = wb['sheet1']

log_headers = ['Correlation Sys ID', 'Number', 'Content Size', 'Result']
log_book = Workbook()
log_sheet = log_book.active
log_sheet.append(log_headers)

user = config.username
pwd = config.password
count = 0
print("Preparing to download ...")
print("===================================================")
print("")
for row in ws.iter_rows(min_row=2):
    try:
        #Current 0 indexed column number for the correlation_sys_id and the records number for the output.xlsx file
        correlation_sys_id = row[46].value
        original_record_number = row[24].value
        
        url = config.URI_Attachments + 'sysparm_sys_id=' + correlation_sys_id + '&sysparm_table='+table
        r = requests.get(url, auth=(user, pwd))
        size = len(r.content)

        if r.status_code != 200:
            log_sheet.append([correlation_sys_id, original_record_number, size, 'Failed - ' + str(r.status_code)])

        if size > 0:
            count +=1
            z = zipfile.ZipFile(io.BytesIO(r.content))
            z.extractall(extract_path + original_record_number)
            print('Finished downloading attachment(s) for: ' + original_record_number)
            log_sheet.append([correlation_sys_id, original_record_number, size , 'Success'])
    except Exception as e:
        log_sheet.append([correlation_sys_id, original_record_number, size, 'Error - ' + str(e)])
        
log_book.save('download-log.xlsx')
print("")
print("=========================RESULTS==========================")
print("DONE! Extracted attachments from " + str(count) + ' records.')
print("Please view the download-log file for additional information")
